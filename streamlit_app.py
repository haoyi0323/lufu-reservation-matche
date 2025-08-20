import streamlit as st
import pandas as pd
import re
import os
from datetime import datetime
import io
import plotly.express as px
import plotly.graph_objects as go
from collections import Counter

class ReservationMatcherWeb:
    def __init__(self):
        self.meituan_file = None
        self.reservation_file = None
        self.merged_df = pd.DataFrame()
        self.original_df = pd.DataFrame()
    
    def process_new_format_reservation(self, df):
        """处理新格式的预定表（8月预定.xls格式）"""
        if df.empty:
            return df
            
        # 检测是否为新格式：第一行包含'包厢'和日期信息，且有很多Unnamed列
        first_row = df.iloc[0] if len(df) > 0 else pd.Series()
        unnamed_cols = [col for col in df.columns if 'Unnamed' in str(col)]
        
        # 新格式特征：第一行有'包厢'，且有多个Unnamed列
        is_new_format = (
            len(unnamed_cols) >= 5 and 
            len(df) > 0 and 
            pd.notna(first_row.iloc[0]) and 
            str(first_row.iloc[0]) == '包厢'
        )
        
        if is_new_format:
            # 处理新格式
            # 跳过第0行（表头行），从第1行开始读取数据
            data_df = df.iloc[1:].copy()
            
            # 重新定义列名
            new_columns = {
                0: '包厢',
                1: '市别', 
                2: '预订时间',
                3: '姓名',
                4: '人数',
                5: '联系电话',
                6: '预订人',
                7: '经手人',
                8: '备注'
            }
            
            # 重命名列
            column_mapping = {}
            for i, col in enumerate(data_df.columns):
                if i in new_columns:
                    column_mapping[col] = new_columns[i]
            
            data_df = data_df.rename(columns=column_mapping)
            
            # 过滤掉空行和无效数据
            data_df = data_df[data_df['包厢'].notna()]
            data_df = data_df[data_df['包厢'] != '晚市']  # 过滤掉分隔行
            
            # 过滤掉总结行（包含"合计"的行）
            data_df = data_df[~data_df['包厢'].astype(str).str.contains('合计', na=False)]
            data_df = data_df[~data_df['包厢'].astype(str).str.contains('总计', na=False)]
            data_df = data_df[~data_df['包厢'].astype(str).str.contains('小计', na=False)]
            
            # 从表头行提取日期信息
            header_info = first_row.iloc[1] if len(first_row) > 1 else None
            if pd.notna(header_info) and '月' in str(header_info):
                # 解析日期信息，如"8月1号 星期五"
                date_str = str(header_info)
                try:
                    # 提取月份和日期
                    import re
                    match = re.search(r'(\d+)月(\d+)号', date_str)
                    if match:
                        month = int(match.group(1))
                        day = int(match.group(2))
                        # 假设是当前年份
                        current_year = datetime.now().year
                        date_obj = datetime(current_year, month, day)
                        data_df['日期'] = date_obj
                except:
                    # 如果解析失败，使用当前日期
                    data_df['日期'] = datetime.now().date()
            else:
                data_df['日期'] = datetime.now().date()
            
            # 处理预订时间字段（新格式中可能是time对象）
            if '预订时间' in data_df.columns:
                def convert_time_format(time_val):
                    if pd.isna(time_val):
                        return None
                    try:
                        # 如果是time对象，转换为字符串
                        if hasattr(time_val, 'strftime'):
                            return time_val.strftime('%H:%M')
                        # 如果是字符串，直接返回
                        elif isinstance(time_val, str):
                            return time_val
                        # 其他情况转换为字符串
                        else:
                            return str(time_val)
                    except:
                        return str(time_val) if time_val is not None else None
                
                data_df['预订时间'] = data_df['预订时间'].apply(convert_time_format)
            
            return data_df
        else:
            # 原格式，直接返回
            return df
        
    def smart_table_match(self, reservation_table, meituan_table):
        """智能桌牌号匹配函数 - 支持新格式包厢名称"""
        # 提取数字部分
        def extract_numbers(table_str):
            if pd.isna(table_str):
                return None
            numbers = re.findall(r'\d+', str(table_str))
            return ''.join(numbers) if numbers else None
        
        # 提取包厢名称关键词
        def extract_room_keywords(table_str):
            if pd.isna(table_str):
                return set()
            table_str = str(table_str).lower()
            # 新格式包厢关键词
            room_keywords = ['福禄', '喜乐', '大厅', '包厢', '雅间']
            found_keywords = set()
            for keyword in room_keywords:
                if keyword in table_str:
                    found_keywords.add(keyword)
            return found_keywords
        
        # 判断是否为外卖订单
        def is_takeout(table_str):
            if pd.isna(table_str):
                return False
            table_str = str(table_str).lower()
            takeout_keywords = ['外卖', 'takeout', '配送', '打包']
            return any(keyword in table_str for keyword in takeout_keywords)
        
        # 完全匹配（最高优先级）
        if str(reservation_table) == str(meituan_table):
            return True, "完全匹配"
        
        # 包厢名称 + 数字匹配（新格式支持）
        res_keywords = extract_room_keywords(reservation_table)
        mt_keywords = extract_room_keywords(meituan_table)
        res_numbers = extract_numbers(reservation_table)
        mt_numbers = extract_numbers(meituan_table)
        
        # 如果包厢关键词和数字都匹配
        if (res_keywords and mt_keywords and 
            res_keywords.intersection(mt_keywords) and 
            res_numbers and mt_numbers and res_numbers == mt_numbers):
            if is_takeout(meituan_table):
                return True, "包厢外卖匹配"
            else:
                return True, "包厢匹配"
        
        # 数字部分匹配（传统匹配方式）
        if res_numbers and mt_numbers and res_numbers == mt_numbers:
            # 区分外卖和堂食的数字匹配
            if is_takeout(meituan_table):
                return True, "外卖匹配"
            else:
                return True, "数字匹配"
        
        return False, "无匹配"
    
    def show_record_details(self, selected_record, display_df, selected_idx):
        """显示选中记录的详细信息"""
        st.divider()
        st.subheader("🔍 记录详情")
        
        # 创建两列布局
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("### 📋 预订信息")
            reservation_info = {
                "日期": selected_record.get('日期', ''),
                "桌牌号": selected_record.get('桌牌号', ''),
                "预订人": selected_record.get('预订人', ''),
                "市别": selected_record.get('市别', ''),
                "匹配状态": selected_record.get('匹配状态', ''),
                "匹配类型": selected_record.get('匹配类型', '')
            }
            
            for key, value in reservation_info.items():
                st.text(f"{key}: {value}")
        
        with col2:
            st.markdown("### 🛒 美团订单信息")
            if selected_record.get('匹配状态') == '已匹配':
                meituan_info = {
                    "下单时间": selected_record.get('下单时间', ''),
                    "桌牌号": selected_record.get('桌牌号', ''),
                    "支付合计": selected_record.get('支付合计', ''),
                    "结账方式": selected_record.get('结账方式', ''),
                    "下单时间格式化": selected_record.get('下单时间_格式化', '')
                }
                
                for key, value in meituan_info.items():
                    st.text(f"{key}: {value}")
                
                # 移除匹配按钮
                st.markdown("---")
                if st.button("❌ 移除此匹配", key=f"remove_{selected_idx}", type="secondary"):
                    self.remove_match(selected_record, selected_idx)
                    st.rerun()
            else:
                st.info("此记录未匹配到美团订单")
    
    def remove_match(self, selected_record, selected_idx):
        """移除匹配记录"""
        try:
            # 在merged_df中找到对应记录并移除匹配信息
            mask = (
                (self.merged_df['日期'] == selected_record['日期']) &
                (self.merged_df['桌牌号'] == selected_record['桌牌号']) &
                (self.merged_df['预订人'] == selected_record['预订人']) &
                (self.merged_df['市别'] == selected_record['市别'])
            )
            
            # 更新匹配状态和相关字段
            self.merged_df.loc[mask, '匹配状态'] = '未匹配'
            self.merged_df.loc[mask, '匹配类型'] = '未匹配'
            self.merged_df.loc[mask, '支付合计'] = None
            self.merged_df.loc[mask, '下单时间'] = None
            self.merged_df.loc[mask, '下单时间_格式化'] = None
            self.merged_df.loc[mask, '结账方式'] = None
            
            st.success("✅ 已成功移除匹配")
            
        except Exception as e:
            st.error(f"❌ 移除匹配失败: {str(e)}")
        
    def load_files(self):
        """现代化文件上传界面"""
        # 美团订单文件上传区域
        st.markdown("""
        <div style='background: linear-gradient(135deg, rgba(59, 130, 246, 0.1), rgba(139, 92, 246, 0.1)); 
                    border-radius: 16px; 
                    padding: 1.5rem; 
                    margin-bottom: 1.5rem; 
                    border: 1px solid rgba(59, 130, 246, 0.2);'>
            <h4 style='margin: 0 0 1rem 0; color: #3b82f6; font-weight: 600;'>
                📊 美团订单文件
            </h4>
            <p style='margin: 0; color: #64748b; font-size: 0.9rem;'>
                支持 .xlsx 和 .xls 格式，请确保文件包含营业日期和桌牌号列
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        meituan_uploaded = st.file_uploader(
            "拖拽文件到此处或点击选择美团订单Excel文件", 
            type=['xlsx', 'xls'],
            key="meituan",
            help="支持的格式：Excel (.xlsx, .xls)，文件大小限制：200MB"
        )
        
        if meituan_uploaded:
            try:
                # 尝试不同的header设置来读取美团文件
                meituan_df = None
                for header_row in [2, 1, 0, None]:
                    try:
                        temp_df = pd.read_excel(meituan_uploaded, header=header_row)
                        # 检查是否包含关键列
                        if any('营业日期' in str(col) for col in temp_df.columns) and \
                           any('桌牌号' in str(col) for col in temp_df.columns):
                            meituan_df = temp_df
                            break
                    except:
                        continue
                
                if meituan_df is None:
                    st.error("无法识别美团文件格式，请检查文件是否正确")
                    return
                
                self.meituan_file = meituan_df
                
                # 清理数据：移除完全空的列和行
                self.meituan_file = self.meituan_file.dropna(how='all', axis=1)  # 删除全空列
                self.meituan_file = self.meituan_file.dropna(how='all', axis=0)  # 删除全空行
                
                # 转换所有列为字符串类型以避免类型冲突
                for col in self.meituan_file.columns:
                    if self.meituan_file[col].dtype == 'object':
                        self.meituan_file[col] = self.meituan_file[col].astype(str)
                    
                # 智能检测列名
                date_col = None
                table_col = None
                customer_col = None
                
                for col in self.meituan_file.columns:
                    if any(keyword in str(col) for keyword in ['营业日期', '日期', 'date']):
                        date_col = col
                    if any(keyword in str(col) for keyword in ['桌牌号', '桌号', '台号']):
                        table_col = col
                    if any(keyword in str(col) for keyword in ['客户', '姓名', '顾客']):
                        customer_col = col
                
                missing_cols = []
                if not date_col: missing_cols.append('日期相关列')
                if not table_col: missing_cols.append('桌牌号相关列')
                
                if missing_cols:
                    st.error(f"❌ 缺少必要列: {', '.join(missing_cols)}")
                else:
                    # 现代化成功提示
                    st.markdown(f"""
                    <div style='background: linear-gradient(135deg, rgba(16, 185, 129, 0.1), rgba(5, 150, 105, 0.1)); 
                                border-radius: 12px; 
                                padding: 1rem 1.5rem; 
                                border-left: 4px solid #10b981; 
                                margin: 1rem 0;'>
                        <div style='display: flex; align-items: center; gap: 0.5rem;'>
                            <span style='font-size: 1.2rem;'>✅</span>
                            <strong style='color: #059669;'>美团文件加载成功！</strong>
                        </div>
                        <p style='margin: 0.5rem 0 0 0; color: #064e3b; font-size: 0.9rem;'>
                            已成功加载 <strong>{len(self.meituan_file)}</strong> 条记录，检测到日期列：<strong>{date_col}</strong>，桌牌号列：<strong>{table_col}</strong>
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    with st.expander("👀 预览美团数据", expanded=False):
                        # 创建显示用的DataFrame副本
                        display_df = self.meituan_file.copy()
                        
                        # 现代化表格样式
                        st.markdown("""
                        <style>
                        .stDataFrame {
                            background: linear-gradient(135deg, rgba(255,255,255,0.95), rgba(248,250,252,0.95));
                            backdrop-filter: blur(10px);
                            border-radius: 12px;
                            overflow: hidden;
                            box-shadow: 0 4px 20px rgba(0, 0, 0, 0.08);
                            border: 1px solid rgba(226, 232, 240, 0.5);
                        }
                        .stDataFrame > div {
                            overflow-x: auto;
                            border-radius: 12px;
                        }
                        .stDataFrame table {
                            border-collapse: separate;
                            border-spacing: 0;
                        }
                        .stDataFrame th {
                            background: linear-gradient(135deg, #3b82f6, #8b5cf6);
                            color: white;
                            font-weight: 600;
                            padding: 12px 16px;
                            border: none;
                            position: sticky;
                            top: 0;
                            z-index: 10;
                        }
                        .stDataFrame td {
                            padding: 10px 16px;
                            border-bottom: 1px solid rgba(226, 232, 240, 0.5);
                            transition: background-color 0.2s ease;
                        }
                        .stDataFrame tr:hover td {
                            background-color: rgba(59, 130, 246, 0.05);
                        }
                        </style>
                        """, unsafe_allow_html=True)
                        
                        # 数据统计信息
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("📊 总记录数", len(display_df))
                        with col2:
                            st.metric("📅 列数", len(display_df.columns))
                        with col3:
                            if date_col and date_col in display_df.columns:
                                unique_dates = display_df[date_col].nunique()
                                st.metric("📆 日期范围", unique_dates)
                        
                        st.dataframe(display_df, use_container_width=True, height=400)
                    
            except Exception as e:
                st.error(f"美团文件加载失败: {str(e)}")
            
            # 分隔线美化
            st.markdown("""
            <div style='height: 1px; 
                        background: linear-gradient(90deg, transparent, rgba(59, 130, 246, 0.3), transparent); 
                        margin: 2rem 0;'></div>
            """, unsafe_allow_html=True)
            
            # 预订记录文件上传区域
            st.markdown("""
            <div style='background: linear-gradient(135deg, rgba(139, 92, 246, 0.1), rgba(59, 130, 246, 0.1)); 
                        border-radius: 16px; 
                        padding: 1.5rem; 
                        margin-bottom: 1.5rem; 
                        border: 1px solid rgba(139, 92, 246, 0.2);'>
                <h4 style='margin: 0 0 1rem 0; color: #8b5cf6; font-weight: 600;'>
                    📋 预订记录文件
                </h4>
                <p style='margin: 0; color: #64748b; font-size: 0.9rem;'>
                    支持多工作表Excel文件，系统将自动合并所有有效数据
                </p>
            </div>
            """, unsafe_allow_html=True)
            
            reservation_uploaded = st.file_uploader(
                "拖拽文件到此处或点击选择预订记录Excel文件", 
                type=['xlsx', 'xls'],
                key="reservation",
                help="支持的格式：Excel (.xlsx, .xls)，可包含多个工作表"
            )
            
            if reservation_uploaded:
                try:
                    # 读取Excel文件的所有工作表
                    excel_file = pd.ExcelFile(reservation_uploaded)
                    all_sheets_data = []
                    
                    # 简化显示信息
                    valid_sheets = 0
                    total_records = 0
                    
                    # 逐个读取每个工作表
                    for sheet_name in excel_file.sheet_names:
                        try:
                            sheet_df = pd.read_excel(reservation_uploaded, sheet_name=sheet_name)
                            
                            # 处理新格式的预定表（检测是否为新格式）
                            sheet_df = self.process_new_format_reservation(sheet_df)
                            
                            # 清理数据：移除完全空的列和行
                            sheet_df = sheet_df.dropna(how='all', axis=1)  # 删除全空列
                            sheet_df = sheet_df.dropna(how='all', axis=0)  # 删除全空行
                            
                            # 如果工作表有数据，添加到列表中
                            if not sheet_df.empty:
                                # 添加工作表名称列用于标识数据来源
                                sheet_df['数据来源工作表'] = sheet_name
                                all_sheets_data.append(sheet_df)
                                valid_sheets += 1
                                total_records += len(sheet_df)
                        except Exception as e:
                            continue  # 静默跳过错误的工作表
                    
                    # 合并所有工作表的数据
                    if all_sheets_data:
                        self.reservation_file = pd.concat(all_sheets_data, ignore_index=True)
                        
                        # 转换所有列为字符串类型以避免类型冲突
                        for col in self.reservation_file.columns:
                            if self.reservation_file[col].dtype == 'object':
                                self.reservation_file[col] = self.reservation_file[col].astype(str)
                        
                        # 现代化成功提示
                        st.markdown(f"""
                        <div style='background: linear-gradient(135deg, rgba(139, 92, 246, 0.1), rgba(124, 58, 237, 0.1)); 
                                    border-radius: 12px; 
                                    padding: 1rem 1.5rem; 
                                    border-left: 4px solid #8b5cf6; 
                                    margin: 1rem 0;'>
                            <div style='display: flex; align-items: center; gap: 0.5rem;'>
                                <span style='font-size: 1.2rem;'>✅</span>
                                <strong style='color: #7c3aed;'>预订文件加载成功！</strong>
                            </div>
                            <p style='margin: 0.5rem 0 0 0; color: #581c87; font-size: 0.9rem;'>
                                已成功处理 <strong>{valid_sheets}</strong> 个工作表，合并 <strong>{len(self.reservation_file)}</strong> 条预订记录
                            </p>
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        st.error("没有找到有效数据")
                        self.reservation_file = pd.DataFrame()
                    
                    with st.expander("👀 预览预订数据", expanded=False):
                        # 创建显示用的DataFrame副本
                        display_df = self.reservation_file.copy()
                        
                        # 现代化表格样式（预订数据用紫色主题）
                        st.markdown("""
                        <style>
                        .reservation-table .stDataFrame {
                            background: linear-gradient(135deg, rgba(255,255,255,0.95), rgba(248,250,252,0.95));
                            backdrop-filter: blur(10px);
                            border-radius: 12px;
                            overflow: hidden;
                            box-shadow: 0 4px 20px rgba(0, 0, 0, 0.08);
                            border: 1px solid rgba(139, 92, 246, 0.3);
                        }
                        .reservation-table .stDataFrame > div {
                            overflow-x: auto;
                            border-radius: 12px;
                        }
                        .reservation-table .stDataFrame th {
                            background: linear-gradient(135deg, #8b5cf6, #7c3aed);
                            color: white;
                            font-weight: 600;
                            padding: 12px 16px;
                            border: none;
                            position: sticky;
                            top: 0;
                            z-index: 10;
                        }
                        .reservation-table .stDataFrame td {
                            padding: 10px 16px;
                            border-bottom: 1px solid rgba(226, 232, 240, 0.5);
                            transition: background-color 0.2s ease;
                        }
                        .reservation-table .stDataFrame tr:hover td {
                            background-color: rgba(139, 92, 246, 0.05);
                        }
                        </style>
                        """, unsafe_allow_html=True)
                        
                        # 数据统计信息
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("📋 总记录数", len(display_df))
                        with col2:
                            st.metric("📊 列数", len(display_df.columns))
                        with col3:
                            if '数据来源工作表' in display_df.columns:
                                unique_sheets = display_df['数据来源工作表'].nunique()
                                st.metric("📄 工作表数", unique_sheets)
                        
                        # 使用容器包装表格以应用特定样式
                        with st.container():
                            st.markdown('<div class="reservation-table">', unsafe_allow_html=True)
                            st.dataframe(display_df, use_container_width=True, height=400)
                            st.markdown('</div>', unsafe_allow_html=True)
                        
                except Exception as e:
                    st.error(f"❌ 预订文件加载失败: {str(e)}")
            

    
    def validate_files(self):
        """验证文件是否已加载"""
        if self.meituan_file is None or self.reservation_file is None:
            return False, "请先上传美团订单文件和预订记录文件"
        
        if self.meituan_file.empty or self.reservation_file.empty:
            return False, "上传的文件为空，请检查文件内容"
        
        return True, "文件验证通过"
    
    def match_data(self):
        """数据匹配核心逻辑 - 使用与桌面版完全相同的匹配算法"""
        try:
            # 读取美团数据 - 使用与桌面版相同的处理方式
            df = self.meituan_file.copy()
            
            # 数据清洗和预处理
            df = df[df['订单状态'] == '已结账']
            df = df[df['营业日期'] != '--']
            
            # 改进的支付金额提取
            def extract_payment(payment_str):
                if pd.isna(payment_str):
                    return None
                try:
                    # 查找所有数字（包括负数和小数）
                    numbers = re.findall(r'-?\d+\.?\d*', str(payment_str))
                    if numbers:
                        return float(numbers[0])
                except (ValueError, IndexError):
                    pass
                return None
                
            df['支付合计'] = df['结账方式'].apply(extract_payment)
            df['营业日期'] = pd.to_datetime(df['营业日期'], errors='coerce')
            df['下单时间'] = pd.to_datetime(df['下单时间'], errors='coerce')
            
            # 根据下单时间判断市别
            def determine_market_period(order_time):
                if pd.isna(order_time):
                    return None
                try:
                    hour = pd.to_datetime(order_time).hour
                    # 午市: 6:00-16:00, 晚市: 16:00-24:00
                    if 6 <= hour < 16:
                        return '午市'
                    elif 16 <= hour <= 23:
                        return '晚市'
                    else:
                        return None  # 非营业时间
                except:
                    return None
            
            df['市别'] = df['下单时间'].apply(determine_market_period)
            
            # 选择需要的列，保留下单时间和结账方式用于显示
            mt_df = df[['营业日期', '桌牌号', '下单时间', '支付合计', '市别', '结账方式']].copy()
            # 过滤掉非营业时间的订单
            mt_df = mt_df[mt_df['市别'].notna()]
            
            # 格式化下单时间为更易读的格式
            try:
                mt_df['下单时间_格式化'] = mt_df['下单时间'].dt.strftime('%H:%M:%S')
            except AttributeError:
                # 如果不是datetime类型，尝试转换后格式化
                mt_df['下单时间'] = pd.to_datetime(mt_df['下单时间'], errors='coerce')
                mt_df['下单时间_格式化'] = mt_df['下单时间'].dt.strftime('%H:%M:%S')
            
            # 提取下单时间的日期部分用于匹配
            mt_df['下单日期'] = mt_df['下单时间'].dt.date
            
            # 读取预订数据
            merged_all = pd.DataFrame()
            
            # 处理预订数据 - 支持多工作表
            if hasattr(self.reservation_file, 'sheet_names'):
                # 如果是ExcelFile对象，处理多个工作表
                for sheet_name in self.reservation_file.sheet_names:
                    try:
                        day_df = pd.read_excel(self.reservation_file, sheet_name=sheet_name)
                        
                        # 检查必要的列是否存在（兼容新旧格式）
                        # 新格式：姓名、预订人
                        # 旧格式：姓名、预订人 或 客户姓名、预订人
                        has_name_col = '姓名' in day_df.columns or '客户姓名' in day_df.columns
                        has_booker_col = '预订人' in day_df.columns
                        
                        if not (has_name_col and has_booker_col):
                            continue
                            
                        # 数据清洗（兼容新旧格式）
                        name_col = '姓名' if '姓名' in day_df.columns else '客户姓名'
                        day_df = day_df[day_df[name_col].notna() & day_df['预订人'].notna()]
                        
                        # 预订人姓名标准化处理
                        def standardize_name(name):
                            if pd.isna(name):
                                return name
                            name_str = str(name).strip()
                            # 处理同义词
                            if name_str in ['平和', '平哥']:
                                return '平和'
                            # 处理刘霞和刘的映射
                            if name_str in ['刘霞', '刘']:
                                return '刘霞'
                            # 处理周和周思玗的映射
                            if name_str in ['周', '周思玗']:
                                return '周思玗'
                            # 处理大小写统一（sk -> SK）
                            if name_str.lower() == 'sk':
                                return 'SK'
                            return name_str
                        
                        day_df['预订人'] = day_df['预订人'].apply(standardize_name)
                        
                        # 选择和重命名列（兼容新旧格式）
                        # 新格式可能的列：日期、市别、包厢、姓名、预订人、人数、时间、客户类型
                        # 旧格式可能的列：日期、市别、包厢、客户姓名、预订人、经手人
                        available_cols = ['日期', '市别', '包厢', '姓名', '客户姓名', '预订人', '经手人', '人数', '时间', '客户类型']
                        existing_cols = [col for col in available_cols if col in day_df.columns]
                        day_df = day_df[existing_cols].copy()
                        
                        # 标准化列名（统一为旧格式的列名以保持兼容性）
                        col_mapping = {
                            '包厢': '桌牌号', 
                            '姓名': '客户姓名',  # 新格式的姓名映射为客户姓名
                            '客户姓名': '客户姓名'  # 旧格式保持不变
                        }
                        day_df.rename(columns=col_mapping, inplace=True)
                        
                        # 处理日期
                        if '日期' in day_df.columns:
                            day_df['日期'] = pd.to_datetime(
                                day_df['日期'].astype(str).str.split().str[0], 
                                errors='coerce'
                            )
                        
                        # 合并数据 - 改进的匹配逻辑
                        if '日期' in day_df.columns and '桌牌号' in day_df.columns and '市别' in day_df.columns:
                            # 为每个预订记录找到最佳匹配的美团订单
                            merged_records = []
                            
                            for _, reservation in day_df.iterrows():
                                # 找到同一日期、市别的所有美团订单，然后使用智能桌牌号匹配
                                reservation_date = reservation['日期'].date() if hasattr(reservation['日期'], 'date') else reservation['日期']
                                
                                # 先按日期和市别筛选
                                candidate_orders = mt_df[
                                    (mt_df['下单日期'] == reservation_date) &
                                    (mt_df['市别'] == reservation['市别'])
                                ].copy()
                                
                                # 使用智能匹配找到桌牌号匹配的订单
                                matching_orders = []
                                match_info = []
                                
                                for _, order in candidate_orders.iterrows():
                                    is_match, match_type = self.smart_table_match(
                                        reservation['桌牌号'], 
                                        order['桌牌号']
                                    )
                                    if is_match:
                                        matching_orders.append(order)
                                        match_info.append(match_type)
                                
                                matching_orders = pd.DataFrame(matching_orders) if matching_orders else pd.DataFrame()
                                
                                if not matching_orders.empty:
                                    # 为每个匹配的订单创建独立记录
                                    for idx, (_, order) in enumerate(matching_orders.iterrows()):
                                        merged_record = reservation.copy()
                                        merged_record['支付合计'] = order['支付合计']
                                        merged_record['下单时间'] = order['下单时间']
                                        merged_record['下单时间_格式化'] = order['下单时间_格式化']
                                        merged_record['结账方式'] = order['结账方式']
                                        merged_record['匹配类型'] = match_info[idx] if idx < len(match_info) else '未知'
                                        merged_records.append(merged_record)
                                else:
                                    # 没有匹配的订单
                                    merged_record = reservation.copy()
                                    merged_record['支付合计'] = None
                                    merged_record['下单时间'] = None
                                    merged_record['下单时间_格式化'] = None
                                    merged_record['结账方式'] = None
                                    merged_record['匹配类型'] = '未匹配'
                                    merged_records.append(merged_record)
                            
                            if merged_records:
                                merged = pd.DataFrame(merged_records)
                                merged_all = pd.concat([merged_all, merged], ignore_index=True)
                                
                    except Exception as e:
                        continue
            else:
                # 如果是单个DataFrame，直接处理
                day_df = self.reservation_file.copy()
                
                # 检查必要的列是否存在 - 兼容新旧格式
                name_col = None
                if '姓名' in day_df.columns:
                    name_col = '姓名'
                elif '客户姓名' in day_df.columns:
                    name_col = '客户姓名'
                
                if name_col is None or '预订人' not in day_df.columns:
                    return False, f"预订文件缺少必要列: 需要'姓名'或'客户姓名'列以及'预订人'列"
                    
                # 数据清洗
                day_df = day_df[day_df[name_col].notna() & day_df['预订人'].notna()]
                
                # 选择和重命名列 - 兼容新旧格式
                available_cols = ['日期', '市别', '包厢', '桌牌号', name_col, '预订人', '经手人', '预订时间']
                existing_cols = [col for col in available_cols if col in day_df.columns]
                day_df = day_df[existing_cols].copy()
                
                # 标准化列名 - 统一映射到旧格式列名
                col_mapping = {
                    '包厢': '桌牌号',
                    '姓名': '客户姓名'  # 新格式的'姓名'映射为'客户姓名'
                }
                # 如果已经是'客户姓名'列，则不需要重命名
                if name_col == '客户姓名':
                    col_mapping.pop('姓名', None)
                    
                day_df.rename(columns=col_mapping, inplace=True)
                
                # 处理日期
                if '日期' in day_df.columns:
                    day_df['日期'] = pd.to_datetime(
                        day_df['日期'].astype(str).str.split().str[0], 
                        errors='coerce'
                    )
                
                # 合并数据 - 改进的匹配逻辑
                if '日期' in day_df.columns and '桌牌号' in day_df.columns and '市别' in day_df.columns:
                    # 为每个预订记录找到最佳匹配的美团订单
                    merged_records = []
                    
                    for _, reservation in day_df.iterrows():
                        # 找到同一日期、市别的所有美团订单，然后使用智能桌牌号匹配
                        reservation_date = reservation['日期'].date() if hasattr(reservation['日期'], 'date') else reservation['日期']
                        
                        # 先按日期和市别筛选
                        candidate_orders = mt_df[
                            (mt_df['下单日期'] == reservation_date) &
                            (mt_df['市别'] == reservation['市别'])
                        ].copy()
                        
                        # 使用智能匹配找到桌牌号匹配的订单
                        matching_orders = []
                        match_info = []
                        
                        for _, order in candidate_orders.iterrows():
                            is_match, match_type = self.smart_table_match(
                                reservation['桌牌号'], 
                                order['桌牌号']
                            )
                            if is_match:
                                matching_orders.append(order)
                                match_info.append(match_type)
                        
                        matching_orders = pd.DataFrame(matching_orders) if matching_orders else pd.DataFrame()
                        
                        if not matching_orders.empty:
                            # 为每个匹配的订单创建独立记录
                            for idx, (_, order) in enumerate(matching_orders.iterrows()):
                                merged_record = reservation.copy()
                                merged_record['支付合计'] = order['支付合计']
                                merged_record['下单时间'] = order['下单时间']
                                merged_record['下单时间_格式化'] = order['下单时间_格式化']
                                merged_record['结账方式'] = order['结账方式']
                                merged_record['匹配类型'] = match_info[idx] if idx < len(match_info) else '未知'
                                merged_records.append(merged_record)
                        else:
                            # 没有匹配的订单
                            merged_record = reservation.copy()
                            merged_record['支付合计'] = None
                            merged_record['下单时间'] = None
                            merged_record['下单时间_格式化'] = None
                            merged_record['结账方式'] = None
                            merged_record['匹配类型'] = '未匹配'
                            merged_records.append(merged_record)
                    
                    if merged_records:
                        merged_all = pd.DataFrame(merged_records)
            
            # 数据后处理
            if not merged_all.empty:
                # 添加匹配状态列
                merged_all['匹配状态'] = merged_all['支付合计'].apply(
                    lambda x: '已匹配' if pd.notna(x) else '未匹配'
                )
                
                # 格式化数据
                if '支付合计' in merged_all.columns:
                    merged_all['支付合计'] = merged_all['支付合计'].apply(
                        lambda x: f"{x:.2f}" if pd.notna(x) else ""
                    )
                    
                # 排序
                sort_cols = []
                if '日期' in merged_all.columns:
                    sort_cols.append('日期')
                if '桌牌号' in merged_all.columns:
                    sort_cols.append('桌牌号')
                if sort_cols:
                    merged_all.sort_values(sort_cols, inplace=True, ignore_index=True)
            
            self.merged_df = merged_all
            self.original_df = merged_all.copy()  # 保存原始数据
            
            # 显示统计信息
            total_records = len(self.merged_df)
            matched_records = len(self.merged_df[self.merged_df['匹配状态'] == '已匹配']) if '匹配状态' in self.merged_df.columns else 0
            
            return True, f"匹配完成！总记录: {total_records}, 已匹配: {matched_records}, 未匹配: {total_records - matched_records}"
            
        except Exception as e:
            return False, f"匹配失败: {str(e)}"
    
    def display_results(self):
        """显示匹配结果"""
        if self.merged_df.empty:
            st.warning("暂无数据")
            return
        
        # 现代化匹配统计信息展示
        if '匹配类型' in self.merged_df.columns:
            st.markdown("""
            <div style='background: linear-gradient(135deg, rgba(99, 102, 241, 0.1), rgba(168, 85, 247, 0.1)); 
                        border-radius: 16px; 
                        padding: 1.5rem; 
                        margin: 1rem 0; 
                        border: 1px solid rgba(99, 102, 241, 0.2);'>
                <h3 style='color: #4f46e5; margin: 0 0 1rem 0; font-weight: 600; display: flex; align-items: center; gap: 0.5rem;'>
                    📊 智能匹配统计概览
                </h3>
            </div>
            """, unsafe_allow_html=True)
            
            match_stats = self.merged_df['匹配类型'].value_counts()
            
            # 添加现代化统计卡片的CSS样式
            st.markdown("""
            <style>
            .metric-card {
                background: linear-gradient(135deg, rgba(255, 255, 255, 0.9), rgba(248, 250, 252, 0.9));
                border-radius: 12px;
                padding: 1.2rem;
                margin: 0.5rem 0;
                border: 1px solid rgba(226, 232, 240, 0.8);
                box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
                transition: all 0.3s ease;
                text-align: center;
            }
            .metric-card:hover {
                transform: translateY(-2px);
                box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.1), 0 4px 6px -2px rgba(0, 0, 0, 0.05);
            }
            .metric-value {
                font-size: 2rem;
                font-weight: 700;
                margin: 0.5rem 0;
            }
            .metric-label {
                font-size: 0.9rem;
                color: #64748b;
                font-weight: 500;
            }
            .metric-complete { color: #059669; }
            .metric-room { color: #0891b2; }
            .metric-number { color: #7c3aed; }
            .metric-unmatch { color: #dc2626; }
            .metric-takeout { color: #ea580c; }
            .metric-room-takeout { color: #c2410c; }
            .metric-matched { color: #16a34a; }
            .metric-rate { color: #2563eb; }
            </style>
            """, unsafe_allow_html=True)
            
            # 第一行：主要匹配类型
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                complete_match = match_stats.get('完全匹配', 0)
                st.markdown(f"""
                <div class='metric-card'>
                    <div class='metric-value metric-complete'>🎯 {complete_match}</div>
                    <div class='metric-label'>完全匹配</div>
                    <div style='font-size: 0.8rem; color: #94a3b8; margin-top: 0.3rem;'>桌牌号完全相同</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                room_match = match_stats.get('包厢匹配', 0)
                st.markdown(f"""
                <div class='metric-card'>
                    <div class='metric-value metric-room'>🏠 {room_match}</div>
                    <div class='metric-label'>包厢匹配</div>
                    <div style='font-size: 0.8rem; color: #94a3b8; margin-top: 0.3rem;'>包厢名称数字匹配</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                number_match = match_stats.get('数字匹配', 0)
                st.markdown(f"""
                <div class='metric-card'>
                    <div class='metric-value metric-number'>🔢 {number_match}</div>
                    <div class='metric-label'>数字匹配</div>
                    <div style='font-size: 0.8rem; color: #94a3b8; margin-top: 0.3rem;'>桌牌号数字相同</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col4:
                no_match = match_stats.get('未匹配', 0)
                st.markdown(f"""
                <div class='metric-card'>
                    <div class='metric-value metric-unmatch'>❌ {no_match}</div>
                    <div class='metric-label'>未匹配</div>
                    <div style='font-size: 0.8rem; color: #94a3b8; margin-top: 0.3rem;'>未找到对应订单</div>
                </div>
                """, unsafe_allow_html=True)
            
            # 第二行：外卖匹配类型和总体统计
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                takeout_match = match_stats.get('外卖匹配', 0)
                st.markdown(f"""
                <div class='metric-card'>
                    <div class='metric-value metric-takeout'>🚚 {takeout_match}</div>
                    <div class='metric-label'>外卖匹配</div>
                    <div style='font-size: 0.8rem; color: #94a3b8; margin-top: 0.3rem;'>预订改为外卖</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                room_takeout_match = match_stats.get('包厢外卖匹配', 0)
                st.markdown(f"""
                <div class='metric-card'>
                    <div class='metric-value metric-room-takeout'>🏠🚚 {room_takeout_match}</div>
                    <div class='metric-label'>包厢外卖</div>
                    <div style='font-size: 0.8rem; color: #94a3b8; margin-top: 0.3rem;'>包厢改为外卖</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                total_records = len(self.merged_df)
                matched_records = total_records - no_match
                st.markdown(f"""
                <div class='metric-card'>
                    <div class='metric-value metric-matched'>✅ {matched_records}</div>
                    <div class='metric-label'>已匹配总数</div>
                    <div style='font-size: 0.8rem; color: #94a3b8; margin-top: 0.3rem;'>成功匹配记录</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col4:
                match_rate = round((total_records - no_match) / total_records * 100, 1) if total_records > 0 else 0
                st.markdown(f"""
                <div class='metric-card'>
                    <div class='metric-value metric-rate'>📈 {match_rate}%</div>
                    <div class='metric-label'>匹配成功率</div>
                    <div style='font-size: 0.8rem; color: #94a3b8; margin-top: 0.3rem;'>总体匹配比例</div>
                </div>
                """, unsafe_allow_html=True)
            
            st.divider()
        
        # 现代化筛选和搜索区域
        st.markdown("""
        <div style='background: linear-gradient(135deg, rgba(248, 250, 252, 0.8), rgba(241, 245, 249, 0.8)); 
                    border-radius: 12px; 
                    padding: 1.5rem; 
                    margin: 1rem 0; 
                    border: 1px solid rgba(226, 232, 240, 0.6);'>
            <h4 style='color: #475569; margin: 0 0 1rem 0; font-weight: 600; display: flex; align-items: center; gap: 0.5rem;'>
                🔍 数据筛选与搜索
            </h4>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**📊 显示范围**")
            filter_option = st.selectbox(
                "选择要显示的数据类型",
                ["全部记录", "已匹配记录", "未匹配记录"],
                help="选择要查看的数据范围"
            )
            # 保存筛选条件到session_state
            st.session_state.filter_option = filter_option
        
        with col2:
            st.markdown("**👤 预订人搜索**")
            search_keyword = st.text_input(
                "输入预订人姓名", 
                placeholder="🔍 输入预订人姓名进行精确搜索...",
                help="支持模糊搜索，输入部分姓名即可"
            )
            # 保存搜索关键词到session_state
            st.session_state.search_keyword = search_keyword
        
        # 应用筛选
        display_df = self.merged_df.copy()
        
        if filter_option == "已匹配记录":
            display_df = display_df[display_df['匹配状态'] == '已匹配']
        elif filter_option == "未匹配记录":
            display_df = display_df[display_df['匹配状态'] == '未匹配']
        
        if search_keyword:
            if '预订人' in display_df.columns:
                # 标准化搜索关键词
                def standardize_search_keyword(keyword):
                    keyword = keyword.strip()
                    if keyword in ['平和', '平哥']:
                        return ['平和', '平哥']  # 返回所有同义词
                    elif keyword in ['刘霞', '刘']:
                        return ['刘霞', '刘']  # 返回刘霞和刘的所有变体
                    elif keyword in ['周', '周思玗']:
                        return ['周', '周思玗']  # 返回周和周思玗的所有变体
                    elif keyword.lower() == 'sk':
                        return ['SK', 'sk', 'Sk', 'sK']  # 返回所有大小写变体
                    else:
                        return [keyword]
                
                search_terms = standardize_search_keyword(search_keyword)
                # 创建搜索条件，匹配任何一个同义词
                search_condition = False
                for term in search_terms:
                    search_condition |= display_df['预订人'].astype(str).str.contains(term, case=False, na=False)
                display_df = display_df[search_condition]
        
        # 现代化数据表格展示
        st.markdown(f"""
        <div style='background: linear-gradient(135deg, rgba(59, 130, 246, 0.1), rgba(147, 51, 234, 0.1)); 
                    border-radius: 12px; 
                    padding: 1.5rem; 
                    margin: 1rem 0; 
                    border: 1px solid rgba(59, 130, 246, 0.2);'>
            <h4 style='color: #3730a3; margin: 0; font-weight: 600; display: flex; align-items: center; gap: 0.5rem;'>
                📋 智能匹配数据表格 
                <span style='background: linear-gradient(135deg, #3b82f6, #8b5cf6); 
                            color: white; 
                            padding: 0.2rem 0.8rem; 
                            border-radius: 20px; 
                            font-size: 0.8rem; 
                            font-weight: 500;'>
                    {len(display_df)} 条记录
                </span>
            </h4>
        </div>
        """, unsafe_allow_html=True)
        
        if not display_df.empty:
            # 配置核心列显示（简化信息）
            columns_to_show = ['日期', '桌牌号', '预订人', '市别', '匹配状态', '匹配类型']
            available_columns = [col for col in columns_to_show if col in display_df.columns]
            
            # 创建显示用的DataFrame副本并处理数据类型
            table_df = display_df[available_columns].copy()
            
            # 格式化显示
            for col in table_df.columns:
                if col == '匹配状态':
                    table_df[col] = table_df[col].apply(lambda x: '✅已匹配' if str(x) == '已匹配' else '❌未匹配')
                elif col == '匹配类型':
                    # 为匹配类型添加图标
                    type_icons = {
                        '完全匹配': '🎯完全匹配',
                        '包厢匹配': '🏠包厢匹配', 
                        '数字匹配': '🔢数字匹配',
                        '外卖匹配': '🚚外卖匹配',
                        '包厢外卖匹配': '🏠🚚包厢外卖',
                        '未匹配': '❌未匹配'
                    }
                    table_df[col] = table_df[col].apply(lambda x: type_icons.get(str(x), str(x)) if pd.notna(x) else '')
                else:
                    table_df[col] = table_df[col].astype(str).replace('nan', '')
            
            # 重命名列标题使其更简洁
            column_rename = {
                '日期': '📅 日期',
                '桌牌号': '🪑 桌号', 
                '预订人': '👤 预订人',
                '市别': '🏪 市别',
                '匹配状态': '📊 状态',
                '匹配类型': '🔍 匹配类型'
            }
            table_df = table_df.rename(columns=column_rename)
            
            # 添加现代化表格样式
            st.markdown("""
            <style>
            .stDataFrame {
                background: linear-gradient(135deg, rgba(255, 255, 255, 0.95), rgba(248, 250, 252, 0.95));
                border-radius: 12px;
                padding: 1rem;
                box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
                border: 1px solid rgba(226, 232, 240, 0.8);
                overflow-x: auto;
            }
            .stDataFrame > div {
                overflow-x: auto;
                border-radius: 8px;
            }
            .stDataFrame table {
                border-collapse: separate;
                border-spacing: 0;
            }
            .stDataFrame th {
                background: linear-gradient(135deg, #f8fafc, #e2e8f0) !important;
                color: #475569 !important;
                font-weight: 600 !important;
                padding: 0.75rem !important;
                border-bottom: 2px solid #cbd5e1 !important;
            }
            .stDataFrame td {
                padding: 0.75rem !important;
                border-bottom: 1px solid #e2e8f0 !important;
            }
            .stDataFrame tr:hover {
                background-color: rgba(59, 130, 246, 0.05) !important;
            }
            </style>
            """, unsafe_allow_html=True)
            
            # 使用可选择的数据表格
            selected_rows = st.dataframe(
                table_df,
                use_container_width=True,
                height=None,  # 移除高度限制，显示所有内容
                on_select="rerun",
                selection_mode="single-row"
            )
            
            # 处理行选择和详情显示
            if selected_rows.selection.rows:
                selected_idx = selected_rows.selection.rows[0]
                if selected_idx < len(display_df):
                    selected_record = display_df.iloc[selected_idx]
                    self.show_record_details(selected_record, display_df, selected_idx)
            
            # 手动匹配功能
            if filter_option == "未匹配记录" and not display_df.empty:
                self.manual_match_interface(display_df)
        else:
            st.info("📝 没有符合条件的记录")
    
    def manual_match_interface(self, unmatched_df):
        """手动匹配界面"""
        st.write("**🔧 手动匹配**")
        
        if unmatched_df.empty:
            return
        
        # 选择要匹配的预订记录（简化显示）
        reservation_options = []
        for idx, row in unmatched_df.iterrows():
            # 格式化日期，只显示年月日
            date_str = 'N/A'
            if pd.notna(row.get('日期')):
                try:
                    if hasattr(row.get('日期'), 'strftime'):
                        date_str = row.get('日期').strftime('%Y-%m-%d')
                    else:
                        date_str = str(row.get('日期')).split(' ')[0]  # 取空格前的日期部分
                except:
                    date_str = str(row.get('日期', 'N/A'))
            
            option_text = f"📅{date_str} | 🪑{row.get('桌牌号', 'N/A')}桌 | 🏪{row.get('市别', 'N/A')} | 👤{row.get('预订人', 'N/A')}"
            reservation_options.append((option_text, idx))
        
        selected_reservation = st.selectbox(
            "选择要匹配的预订记录",
            options=reservation_options,
            format_func=lambda x: x[0]
        )
        
        if selected_reservation and self.meituan_file is not None:
            reservation_idx = selected_reservation[1]
            reservation_record = unmatched_df.loc[reservation_idx]
            
            # 获取相关的美团订单
            reservation_date = reservation_record['日期']
            if hasattr(reservation_date, 'date'):
                reservation_date = reservation_date.date()
            
            # 处理美团数据，确保支付合计字段正确提取
            def extract_payment(payment_str):
                if pd.isna(payment_str):
                    return None
                try:
                    # 查找所有数字（包括负数和小数）
                    numbers = re.findall(r'-?\d+\.?\d*', str(payment_str))
                    if numbers:
                        return float(numbers[0])
                except (ValueError, IndexError):
                    pass
                return None
            
            # 复制美团文件并处理支付合计
            meituan_processed = self.meituan_file.copy()
            
            # 应用与自动匹配相同的数据过滤条件
            if '订单状态' in meituan_processed.columns:
                meituan_processed = meituan_processed[meituan_processed['订单状态'] == '已结账']
            if '营业日期' in meituan_processed.columns:
                meituan_processed = meituan_processed[meituan_processed['营业日期'] != '--']
            
            if '支付合计' not in meituan_processed.columns and '结账方式' in meituan_processed.columns:
                meituan_processed['支付合计'] = meituan_processed['结账方式'].apply(extract_payment)
            
            # 安全地比较日期（使用下单时间的日期进行匹配）
            try:
                if '下单时间' in meituan_processed.columns:
                    # 确保下单时间是datetime类型
                    meituan_dates = pd.to_datetime(meituan_processed['下单时间'], errors='coerce')
                    related_meituan = meituan_processed[
                        meituan_dates.dt.date == reservation_date
                    ]
                else:
                    related_meituan = meituan_processed
            except (AttributeError, TypeError):
                related_meituan = meituan_processed
            
            if related_meituan.empty:
                related_meituan = meituan_processed
            
            st.write("**📋 可选择的美团订单:**")
            st.write("💡 *点击表格中的行来选择美团订单（支持多选，按住Ctrl键可选择多个）*")
            
            # 添加调试信息
            with st.expander("🔍 数据调试信息", expanded=False):
                st.write("**原始美团数据统计:**")
                original_count = len(self.meituan_file) if self.meituan_file is not None else 0
                processed_count = len(meituan_processed)
                st.write(f"- 原始数据行数: {original_count}")
                st.write(f"- 过滤后行数: {processed_count}")
                
                if '订单状态' in self.meituan_file.columns:
                    status_counts = self.meituan_file['订单状态'].value_counts()
                    st.write("**订单状态分布:**")
                    for status, count in status_counts.items():
                        st.write(f"- {status}: {count} 个")
                
                if '营业日期' in self.meituan_file.columns:
                    dash_count = (self.meituan_file['营业日期'] == '--').sum()
                    st.write(f"**营业日期为'--'的记录数:** {dash_count}")
                
                # 显示原始数据的日期范围
                if '下单时间' in self.meituan_file.columns:
                    try:
                        original_dates = pd.to_datetime(self.meituan_file['下单时间'], errors='coerce')
                        original_valid_dates = original_dates.dropna()
                        if len(original_valid_dates) > 0:
                            st.write(f"**原始数据日期范围:** {original_valid_dates.min().date()} 到 {original_valid_dates.max().date()}")
                            
                            # 按日期统计原始数据
                            original_date_counts = original_valid_dates.dt.date.value_counts().sort_index()
                            st.write("**原始数据按日期统计（前10天）:**")
                            for date, count in original_date_counts.head(10).items():
                                st.write(f"- {date}: {count} 个订单")
                    except Exception as e:
                        st.write(f"原始日期分析错误: {e}")
            
            # 添加日期筛选器
            st.write("**🗓️ 日期筛选:**")
            col1, col2 = st.columns([1, 1])
            
            with col1:
                # 获取美团订单中的所有日期
                available_dates = []
                if '下单时间' in meituan_processed.columns:
                    try:
                        meituan_dates = pd.to_datetime(meituan_processed['下单时间'], errors='coerce')
                        unique_dates = meituan_dates.dt.date.dropna().unique()
                        available_dates = sorted([d for d in unique_dates if d is not None])
                    except:
                        pass
                
                if available_dates:
                    # 默认选择预订记录的日期（如果存在）
                    default_date = reservation_date if reservation_date in available_dates else available_dates[0]
                    selected_date = st.selectbox(
                        "选择要查看的日期",
                        options=available_dates,
                        index=available_dates.index(default_date) if default_date in available_dates else 0,
                        format_func=lambda x: x.strftime('%Y-%m-%d (%A)') if x else 'N/A'
                    )
                    
                    # 根据选择的日期重新筛选美团订单
                    try:
                        meituan_dates = pd.to_datetime(meituan_processed['下单时间'], errors='coerce')
                        related_meituan = meituan_processed[
                            meituan_dates.dt.date == selected_date
                        ]
                    except:
                        related_meituan = meituan_processed
                else:
                    st.info("未找到有效的日期信息")
                    related_meituan = meituan_processed
            
            with col2:
                # 显示筛选结果统计
                if not related_meituan.empty:
                    st.metric("当日美团订单数", len(related_meituan))
                else:
                    st.metric("当日美团订单数", 0)
            
            # 显示美团订单详细信息表格（可选择）
            if not related_meituan.empty:
                # 选择要显示的核心列（包含日期、时间、金额）
                display_columns = ['营业日期', '桌牌号', '下单时间', '支付合计', '结账方式']
                available_columns = [col for col in display_columns if col in related_meituan.columns]
                
                if available_columns:
                    meituan_display = related_meituan[available_columns].copy()
                    # 格式化显示
                    for col in meituan_display.columns:
                        if col == '支付合计':
                            meituan_display[col] = meituan_display[col].apply(lambda x: f"¥{x}" if pd.notna(x) and str(x) != 'nan' else '')
                        elif col == '下单时间':
                            # 确保下单时间显示完整的日期时间
                            meituan_display[col] = meituan_display[col].apply(
                                lambda x: pd.to_datetime(x).strftime('%Y-%m-%d %H:%M:%S') if pd.notna(x) and x != '' else ''
                            )
                        else:
                            meituan_display[col] = meituan_display[col].astype(str).replace('nan', '')
                    
                    # 重命名列标题使其更简洁
                    column_rename = {
                        '营业日期': '📅营业日期',
                        '桌牌号': '🪑桌号',
                        '下单时间': '⏰下单时间', 
                        '支付合计': '💰支付合计',
                        '结账方式': '💳结账方式'
                    }
                    meituan_display = meituan_display.rename(columns=column_rename)
                    
                    # 使用可选择的数据框（支持多选）
                    selected_rows = st.dataframe(
                        meituan_display,
                        use_container_width=True,
                        height=200,
                        hide_index=False,
                        selection_mode="multi-row",
                        key="meituan_selector",
                        on_select="rerun"
                    )
                    
                    # 获取选中的行（支持多选）
                    selected_meituan_indices = []
                    if selected_rows and 'selection' in selected_rows and 'rows' in selected_rows['selection']:
                        if selected_rows['selection']['rows']:
                            for selected_row_idx in selected_rows['selection']['rows']:
                                actual_idx = related_meituan.index[selected_row_idx]
                                selected_meituan_indices.append(actual_idx)
                    
                    # 显示选中的订单数量
                    if selected_meituan_indices:
                        st.info(f"已选择 {len(selected_meituan_indices)} 个美团订单")
            
            # 确认匹配按钮
            if st.button("确认匹配", type="primary"):
                if selected_meituan_indices:
                    # 获取原始预订记录
                    original_reservation = self.merged_df.loc[reservation_idx].copy()
                    
                    # 为每个选中的美团订单创建匹配记录
                    new_records = []
                    for i, meituan_idx in enumerate(selected_meituan_indices):
                        meituan_record = related_meituan.loc[meituan_idx]
                        
                        # 创建新的匹配记录
                        new_record = original_reservation.copy()
                        new_record['匹配状态'] = '已匹配'
                        new_record['下单时间'] = str(meituan_record.get('下单时间', ''))
                        new_record['下单时间_格式化'] = str(meituan_record.get('下单时间', ''))
                        new_record['结账方式'] = str(meituan_record.get('结账方式', ''))
                        
                        # 如果是第一个记录，更新原记录；否则添加新记录
                        if i == 0:
                            # 更新原记录
                            for col in new_record.index:
                                self.merged_df.at[reservation_idx, col] = new_record[col]
                        else:
                            # 添加新记录到列表
                            new_records.append(new_record)
                    
                    # 将新记录添加到DataFrame
                    if new_records:
                        new_df = pd.DataFrame(new_records)
                        self.merged_df = pd.concat([self.merged_df, new_df], ignore_index=True)
                    
                    st.success(f"匹配成功！已为 {len(selected_meituan_indices)} 个美团订单创建匹配记录。页面将自动刷新")
                    st.rerun()
                else:
                    st.warning("请先选择要匹配的美团订单")
    
    def export_results(self):
        """导出结果"""
        if self.merged_df.empty:
            st.warning("暂无数据")
            return
        
        # 导出选项
        export_option = st.selectbox(
            "导出选项",
            ["仅搜索", "全部（按时间排列）"]
        )
        
        # 获取当前搜索和筛选条件
        if 'filter_option' not in st.session_state:
            st.session_state.filter_option = "全部记录"
        if 'search_keyword' not in st.session_state:
            st.session_state.search_keyword = ""
        
        # 准备导出数据
        if export_option == "仅搜索":
            # 获取当前显示的搜索结果（只包含匹配成功的）
            export_df = self.get_filtered_data()
            export_df = export_df[export_df['匹配状态'] == '已匹配']  # 只导出匹配成功的
            
            # 根据搜索关键词生成文件名
            search_keyword = getattr(st.session_state, 'search_keyword', "")
            if search_keyword.strip():
                filename_suffix = f"{search_keyword.strip()}（美团匹配清单）"
            else:
                filename_suffix = "搜索结果"
        else:
            # 全部匹配成功的数据，按时间排列
            export_df = self.merged_df[self.merged_df['匹配状态'] == '已匹配'].copy()
            if '日期' in export_df.columns:
                export_df = export_df.sort_values('日期')
            filename_suffix = "全部匹配"
        
        if export_df.empty:
            st.warning("没有匹配成功的数据可导出")
            return
        
        # 准备导出的列
        export_columns = ['下单时间', '预订人', '桌牌号', '支付合计', '结账方式', '匹配类型']
        
        # 检查并选择可用的列
        available_columns = []
        for col in export_columns:
            if col in export_df.columns:
                available_columns.append(col)
            elif col == '预订人' and '客户姓名' in export_df.columns:
                available_columns.append('客户姓名')
                export_df = export_df.rename(columns={'客户姓名': '预订人'})
        
        # 创建导出用的DataFrame
        final_export_df = export_df[available_columns].copy()
        
        # 按日期排序（如果有下单时间列）
        if '下单时间' in final_export_df.columns:
            final_export_df = final_export_df.sort_values('下单时间')
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            final_export_df.to_excel(writer, sheet_name='匹配结果', index=False)
            
            # 获取工作表并设置格式
            worksheet = writer.sheets['匹配结果']
            
            # 设置Excel格式
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            
            # 定义样式
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 设置表头样式
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            # 设置数据行样式
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = center_alignment
                    cell.border = border
            
            # 智能调整列宽
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                # 计算列的最大内容长度
                for cell in column:
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        # 中文字符按2个字符计算宽度
                        char_count = sum(2 if ord(char) > 127 else 1 for char in cell_value)
                        if char_count > max_length:
                            max_length = char_count
                    except:
                        pass
                
                # 根据列内容设置合适的宽度
                if column_letter == 'A':  # 下单时间列
                    adjusted_width = max(22, min(max_length + 4, 28))
                elif column_letter == 'B':  # 预订人列
                    adjusted_width = max(15, min(max_length + 3, 25))
                elif column_letter == 'C':  # 桌牌号列
                    adjusted_width = max(12, min(max_length + 3, 18))
                elif column_letter == 'D':  # 支付合计列
                    adjusted_width = max(15, min(max_length + 3, 22))
                elif column_letter == 'E':  # 结账方式列
                    adjusted_width = max(25, min(max_length + 5, 40))  # 增加结账方式列宽度
                elif column_letter == 'F':  # 匹配类型列
                    adjusted_width = max(12, min(max_length + 3, 18))
                else:
                    adjusted_width = max(15, min(max_length + 3, 35))
                
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # 设置行高
            for row in range(1, worksheet.max_row + 1):
                worksheet.row_dimensions[row].height = 35  # 增加行高以适应多行内容
            
            # 特别处理表头行高
            worksheet.row_dimensions[1].height = 30
        
        excel_data = output.getvalue()
        
        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"匹配结果_{filename_suffix}_{timestamp}.xlsx"
        
        st.download_button(
            label=f"📥 下载Excel ({len(final_export_df)}条记录)",
            data=excel_data,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    
    def get_filtered_data(self):
        """获取当前筛选和搜索后的数据"""
        display_df = self.merged_df.copy()
        
        # 应用筛选（从session_state获取当前筛选条件）
        filter_option = getattr(st.session_state, 'filter_option', "全部记录")
        search_keyword = getattr(st.session_state, 'search_keyword', "")
        
        if filter_option == "已匹配记录":
            display_df = display_df[display_df['匹配状态'] == '已匹配']
        elif filter_option == "未匹配记录":
            display_df = display_df[display_df['匹配状态'] == '未匹配']
        
        if search_keyword:
            if '预订人' in display_df.columns:
                # 标准化搜索关键词
                def standardize_search_keyword(keyword):
                    keyword = keyword.strip()
                    if keyword in ['平和', '平哥']:
                        return ['平和', '平哥']  # 返回所有同义词
                    elif keyword in ['刘霞', '刘']:
                        return ['刘霞', '刘']  # 返回刘霞和刘的所有变体
                    elif keyword in ['周', '周思玗']:
                        return ['周', '周思玗']  # 返回周和周思玗的所有变体
                    elif keyword.lower() == 'sk':
                        return ['SK', 'sk', 'Sk', 'sK']  # 返回所有大小写变体
                    else:
                        return [keyword]
                
                search_terms = standardize_search_keyword(search_keyword)
                # 创建搜索条件，匹配任何一个同义词
                search_condition = False
                for term in search_terms:
                    search_condition |= display_df['预订人'].astype(str).str.contains(term, case=False, na=False)
                display_df = display_df[search_condition]
        
        return display_df
    
    def normalize_customer_name(self, name):
        """标准化预订人姓名"""
        if pd.isna(name) or str(name).strip() == '':
            return None
        
        name = str(name).strip()
        
        # 转换为小写进行比较
        name_lower = name.lower()
        
        # 定义姓名映射规则
        name_mappings = {
            'sk': 'SK',  # sk -> SK
            '平': '平哥',  # 平 -> 平哥
            '平哥': '平哥',  # 平哥保持不变
            '周': '周',  # 周保持不变
        }
        
        # 检查是否需要映射
        for key, value in name_mappings.items():
            if name_lower == key.lower():
                return value
        
        # 如果没有特殊映射，返回原始名称（保持原有大小写）
        return name
    
    def get_standardized_customers(self):
        """获取标准化后的预订人列表"""
        if '预订人' not in self.merged_df.columns:
            return []
        
        # 标准化所有预订人姓名
        standardized_names = self.merged_df['预订人'].apply(self.normalize_customer_name)
        standardized_names = standardized_names.dropna().unique()
        
        return sorted([name for name in standardized_names if name])
    
    def show_data_analysis(self):
        """显示数据分析页面"""
        st.header("📈 预订人数据分析")
        
        if self.merged_df.empty:
            st.warning("暂无数据，请先在'文件处理'标签页中上传文件并进行匹配")
            return
        
        # 创建两列布局
        col1, col2 = st.columns([1, 2])
        
        with col1:
            st.subheader("🔍 预订人搜索")
            
            # 获取标准化后的预订人列表
            if '预订人' in self.merged_df.columns:
                all_customers = self.get_standardized_customers()
                
                # 搜索框
                search_customer = st.selectbox(
                    "选择预订人",
                    options=["请选择..."] + all_customers,
                    key="customer_analysis_search"
                )
                
                # 或者输入搜索
                manual_search = st.text_input(
                    "或手动输入预订人姓名",
                    placeholder="输入预订人姓名进行搜索",
                    key="manual_customer_search"
                )
                
                # 确定最终搜索的客户
                target_customer = None
                if manual_search.strip():
                    target_customer = manual_search.strip()
                elif search_customer != "请选择...":
                    target_customer = search_customer
                
                if target_customer:
                    st.success(f"已选择：{target_customer}")
                    
                    # 分析按钮
                    if st.button("📊 开始分析", type="primary", use_container_width=True):
                        st.session_state.analysis_customer = target_customer
                        st.rerun()
            else:
                st.error("数据中未找到'预订人'字段")
        
        with col2:
            st.subheader("📊 分析结果")
            
            # 检查是否有选择的客户进行分析
            if hasattr(st.session_state, 'analysis_customer') and st.session_state.analysis_customer:
                customer_name = st.session_state.analysis_customer
                
                # 筛选该客户的数据（使用标准化姓名匹配）
                standardized_customer_names = self.merged_df['预订人'].apply(self.normalize_customer_name)
                customer_data = self.merged_df[standardized_customer_names == customer_name]
                
                if customer_data.empty:
                    st.warning(f"未找到预订人'{customer_name}'的相关数据")
                else:
                    # 显示基本统计信息
                    st.markdown(f"### 👤 {customer_name} 的预订统计")
                    
                    # 创建指标卡片
                    metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)
                    
                    with metric_col1:
                        total_orders = len(customer_data)
                        st.metric("总预订次数", total_orders)
                    
                    with metric_col2:
                        matched_orders = len(customer_data[customer_data['匹配状态'] == '已匹配'])
                        st.metric("成功匹配", matched_orders)
                    
                    with metric_col3:
                        if matched_orders > 0:
                            match_rate = round((matched_orders / total_orders) * 100, 1)
                            st.metric("匹配率", f"{match_rate}%")
                        else:
                            st.metric("匹配率", "0%")
                    
                    with metric_col4:
                        # 计算总消费金额（仅匹配成功的订单）
                        matched_data = customer_data[customer_data['匹配状态'] == '已匹配']
                        if not matched_data.empty and '支付合计' in matched_data.columns:
                            # 提取支付金额的数字部分
                            amounts = matched_data['支付合计'].astype(str).str.extract(r'([0-9.]+)').astype(float)
                            total_amount = amounts.sum().iloc[0] if not amounts.empty else 0
                            st.metric("总消费金额", f"¥{total_amount:.2f}")
                        else:
                            st.metric("总消费金额", "¥0.00")
                    
                    st.divider()
                    
                    # 可视化图表
                    chart_col1, chart_col2 = st.columns(2)
                    
                    with chart_col1:
                        # 工作日vs周末分析（仅分析已匹配数据）
                        st.markdown("#### 📅 工作日vs周末分析")
                        if '日期' in customer_data.columns:
                            # 只分析已匹配的数据
                            matched_data = customer_data[customer_data['匹配状态'] == '已匹配'].copy()
                            
                            if not matched_data.empty:
                                # 转换日期格式并分析工作日/周末
                                matched_data['日期'] = pd.to_datetime(matched_data['日期'], errors='coerce')
                                customer_data_copy = matched_data.dropna(subset=['日期'])
                            else:
                                customer_data_copy = pd.DataFrame()  # 空数据框
                            
                            if not customer_data_copy.empty:
                                # 添加星期几列 (0=周一, 6=周日)
                                customer_data_copy['weekday'] = customer_data_copy['日期'].dt.dayofweek
                                customer_data_copy['day_type'] = customer_data_copy['weekday'].apply(
                                    lambda x: '周末' if x >= 5 else '工作日'
                                )
                                
                                # 统计工作日vs周末的预订次数
                                day_type_counts = customer_data_copy['day_type'].value_counts()
                                
                                if not day_type_counts.empty:
                                    fig_daytype = px.bar(
                                        x=day_type_counts.index,
                                        y=day_type_counts.values,
                                        title=f"{customer_name} 的工作日vs周末预订分布（已匹配数据）",
                                        labels={'x': '日期类型', 'y': '预订次数'},
                                        color=day_type_counts.index,
                                        color_discrete_map={
                                            '工作日': '#3b82f6',
                                            '周末': '#f59e0b'
                                        }
                                    )
                                    fig_daytype.update_layout(
                                        height=300,
                                        font=dict(family="Microsoft YaHei, SimHei, sans-serif"),
                                        showlegend=False
                                    )
                                    st.plotly_chart(fig_daytype, use_container_width=True)
                                    
                                    # 显示统计信息
                                    workday_count = day_type_counts.get('工作日', 0)
                                    weekend_count = day_type_counts.get('周末', 0)
                                    total_count = workday_count + weekend_count
                                    
                                    if total_count > 0:
                                        workday_pct = round((workday_count / total_count) * 100, 1)
                                        weekend_pct = round((weekend_count / total_count) * 100, 1)
                                        
                                        st.markdown(f"""
                                        **📊 统计摘要（已匹配数据）：**
                                        - 工作日预订：{workday_count}次 ({workday_pct}%)
                                        - 周末预订：{weekend_count}次 ({weekend_pct}%)
                                        """)
                                else:
                                    st.info("暂无有效的已匹配日期数据进行工作日/周末分析")
                            else:
                                st.info("该预订人暂无已匹配的数据，无法进行工作日/周末分析")
                        else:
                            st.info("数据中缺少日期字段，无法进行工作日/周末分析")
                    
                    with chart_col2:
                        # 预订时间趋势图
                        st.markdown("#### 📅 预订时间趋势")
                        if '日期' in customer_data.columns:
                            # 按日期统计预订次数
                            date_counts = customer_data['日期'].value_counts().sort_index()
                            
                            if not date_counts.empty:
                                 fig_line = px.line(
                                     x=date_counts.index,
                                     y=date_counts.values,
                                     title=f"{customer_name} 的预订时间趋势",
                                     labels={'x': '日期', 'y': '预订次数'}
                                 )
                                 fig_line.update_layout(
                                      height=300,
                                      font=dict(family="Microsoft YaHei, SimHei, sans-serif"),
                                      xaxis=dict(
                                          tickformat='%Y年%m月%d日',
                                          tickangle=45
                                      )
                                  )
                                 st.plotly_chart(fig_line, use_container_width=True)
                            else:
                                st.info("暂无日期数据")
                        else:
                            st.info("数据中未包含日期信息")
                    
                    # 桌牌号偏好分析
                    if '桌牌号' in customer_data.columns:
                        st.markdown("#### 🪑 桌牌号偏好分析")
                        table_counts = customer_data['桌牌号'].value_counts().head(10)
                        
                        if not table_counts.empty:
                             fig_bar = px.bar(
                                 x=table_counts.values,
                                 y=table_counts.index,
                                 orientation='h',
                                 title=f"{customer_name} 的桌牌号偏好 (前10)",
                                 labels={'x': '预订次数', 'y': '桌牌号'}
                             )
                             fig_bar.update_layout(
                                 height=400,
                                 font=dict(family="Microsoft YaHei, SimHei, sans-serif")
                             )
                             st.plotly_chart(fig_bar, use_container_width=True)
                    
                    # 详细数据表格
                    st.markdown("#### 📋 详细预订记录")
                    
                    # 选择要显示的列
                    display_columns = ['日期', '桌牌号', '匹配状态', '匹配类型']
                    if '支付合计' in customer_data.columns:
                        display_columns.append('支付合计')
                    if '下单时间' in customer_data.columns:
                        display_columns.append('下单时间')
                    
                    available_columns = [col for col in display_columns if col in customer_data.columns]
                    
                    if available_columns:
                        display_data = customer_data[available_columns].copy()
                        
                        # 按日期排序
                        if '日期' in display_data.columns:
                            display_data = display_data.sort_values('日期', ascending=False)
                        
                        st.dataframe(
                            display_data,
                            use_container_width=True,
                            hide_index=True
                        )
                        
                        # 导出该客户的数据
                        if st.button(f"📥 导出 {customer_name} 的数据", use_container_width=True):
                            # 创建Excel文件
                            output = io.BytesIO()
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                display_data.to_excel(writer, sheet_name=f'{customer_name}_预订记录', index=False)
                            
                            excel_data = output.getvalue()
                            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                            filename = f"{customer_name}_预订分析_{timestamp}.xlsx"
                            
                            st.download_button(
                                label=f"下载 {customer_name} 的预订数据",
                                data=excel_data,
                                file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    else:
                        st.warning("无可显示的详细数据")
            else:
                st.info("请在左侧选择一个预订人进行分析")
                
                # 显示整体统计概览
                st.markdown("### 📊 整体数据概览")
                
                # 最活跃的预订人Top 10
                if '预订人' in self.merged_df.columns:
                    # 使用标准化后的姓名进行统计
                    standardized_names = self.merged_df['预订人'].apply(self.normalize_customer_name)
                    valid_customers = standardized_names.dropna()
                    top_customers = valid_customers.value_counts().head(10)
                    
                    if not top_customers.empty:
                         st.markdown("#### 🏆 最活跃预订人 (Top 10)")
                         
                         fig_top = px.bar(
                             x=top_customers.values,
                             y=top_customers.index,
                             orientation='h',
                             title="最活跃的预订人排行榜",
                             labels={'x': '预订次数', 'y': '预订人'}
                         )
                         fig_top.update_layout(
                             height=400,
                             font=dict(family="Microsoft YaHei, SimHei, sans-serif")
                         )
                         st.plotly_chart(fig_top, use_container_width=True)

def main():
    st.set_page_config(
        page_title="鹭府预定匹配工具 v2.0",
        page_icon="📊",
        layout="wide",
        initial_sidebar_state="collapsed"
    )
    
    # 现代化UI设计样式
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    /* CSS变量定义统一配色方案 */
    :root {
        --primary-color: #3b82f6;
        --primary-dark: #2563eb;
        --secondary-color: #8b5cf6;
        --secondary-dark: #7c3aed;
        --accent-color: #10b981;
        --accent-dark: #059669;
        --warning-color: #f59e0b;
        --error-color: #ef4444;
        --success-color: #10b981;
        --info-color: #3b82f6;
        --text-primary: #0f172a;
        --text-secondary: #64748b;
        --text-muted: #94a3b8;
        --bg-primary: #ffffff;
        --bg-secondary: #f8fafc;
        --bg-tertiary: #f1f5f9;
        --border-color: #e2e8f0;
        --shadow-sm: 0 1px 2px 0 rgba(0, 0, 0, 0.05);
        --shadow-md: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
        --shadow-lg: 0 10px 15px -3px rgba(0, 0, 0, 0.1);
        --shadow-xl: 0 20px 25px -5px rgba(0, 0, 0, 0.1);
        --gradient-primary: linear-gradient(135deg, var(--primary-color), var(--secondary-color));
        --gradient-secondary: linear-gradient(135deg, #667eea, #764ba2);
        --gradient-bg: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
    }
    
    .main {
        padding: 1rem 2rem;
        max-width: 1400px;
        margin: 0 auto;
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
    }
    
    /* 全局背景渐变 */
    .stApp {
        background: var(--gradient-bg);
    }
    
    /* 侧边栏美化 */
    .css-1d391kg {
        background: var(--gradient-secondary);
        border-radius: 0 20px 20px 0;
    }
    
    /* 标签页现代化设计 */
    .stTabs [data-baseweb="tab-list"] {
        gap: 0.5rem;
        background: rgba(255, 255, 255, 0.9);
        backdrop-filter: blur(10px);
        border-radius: 15px;
        padding: 0.5rem;
        box-shadow: var(--shadow-lg);
        border: 1px solid rgba(255, 255, 255, 0.2);
    }
    .stTabs [data-baseweb="tab"] {
        height: 3.5rem;
        padding: 0 2rem;
        background: transparent;
        border: none;
        border-radius: 10px;
        font-weight: 600;
        color: var(--text-secondary);
        font-size: 15px;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        position: relative;
        overflow: hidden;
    }
    .stTabs [data-baseweb="tab"]:hover {
        background: rgba(59, 130, 246, 0.1);
        color: var(--primary-color);
        transform: translateY(-2px);
    }
    .stTabs [aria-selected="true"] {
        background: var(--gradient-primary);
        color: white;
        box-shadow: 0 4px 20px rgba(59, 130, 246, 0.4);
        transform: translateY(-1px);
    }
    .stTabs [aria-selected="true"]::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        bottom: 0;
        background: linear-gradient(45deg, rgba(255,255,255,0.1), transparent);
        pointer-events: none;
    }
    /* 文件上传器现代化设计 */
    .stFileUploader {
        border: 2px dashed var(--border-color);
        border-radius: 16px;
        padding: 3rem 2rem;
        background: linear-gradient(135deg, rgba(255,255,255,0.9), rgba(248,250,252,0.9));
        backdrop-filter: blur(10px);
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        position: relative;
        overflow: hidden;
    }
    .stFileUploader::before {
        content: '';
        position: absolute;
        top: 0;
        left: -100%;
        width: 100%;
        height: 100%;
        background: linear-gradient(90deg, transparent, rgba(59, 130, 246, 0.1), transparent);
        transition: left 0.5s;
    }
    .stFileUploader:hover {
        border-color: var(--primary-color);
        background: linear-gradient(135deg, rgba(59, 130, 246, 0.05), rgba(139, 92, 246, 0.05));
        transform: translateY(-2px);
        box-shadow: 0 10px 40px rgba(59, 130, 246, 0.15);
    }
    .stFileUploader:hover::before {
        left: 100%;
    }
    
    /* 指标容器美化 */
    .metric-container {
        background: linear-gradient(135deg, rgba(255,255,255,0.95), rgba(248,250,252,0.95));
        backdrop-filter: blur(15px);
        border: 1px solid rgba(255, 255, 255, 0.2);
        border-radius: 16px;
        padding: 2rem;
        margin: 1rem 0;
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
        transition: all 0.3s ease;
    }
    .metric-container:hover {
        transform: translateY(-4px);
        box-shadow: 0 12px 48px rgba(0, 0, 0, 0.15);
    }
    
    /* 警告框现代化 */
    .stAlert {
        border-radius: 12px;
        border: none;
        backdrop-filter: blur(10px);
        box-shadow: var(--shadow-md);
    }
    .stAlert[data-baseweb="notification"][kind="success"] {
        background: linear-gradient(135deg, rgba(236, 253, 245, 0.9), rgba(220, 252, 231, 0.9));
        color: var(--accent-dark);
        border-left: 4px solid var(--success-color);
    }
    .stAlert[data-baseweb="notification"][kind="error"] {
        background: linear-gradient(135deg, rgba(254, 242, 242, 0.9), rgba(252, 231, 231, 0.9));
        color: #dc2626;
        border-left: 4px solid var(--error-color);
    }
    .stAlert[data-baseweb="notification"][kind="warning"] {
        background: linear-gradient(135deg, rgba(255, 251, 235, 0.9), rgba(254, 243, 199, 0.9));
        color: #d97706;
        border-left: 4px solid var(--warning-color);
    }
    .stAlert[data-baseweb="notification"][kind="info"] {
        background: linear-gradient(135deg, rgba(239, 246, 255, 0.9), rgba(219, 234, 254, 0.9));
        color: var(--primary-dark);
        border-left: 4px solid var(--info-color);
    }
    
    /* 数据表格美化 */
    .stDataFrame {
        border: 1px solid rgba(226, 232, 240, 0.5);
        border-radius: 12px;
        overflow: hidden;
        box-shadow: var(--shadow-md);
        backdrop-filter: blur(10px);
    }
    
    /* 标题样式 */
    h1, h2, h3 {
        color: var(--text-primary);
        font-weight: 700;
        background: var(--gradient-secondary);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
    }
    
    /* 按钮现代化设计 */
    .stButton > button {
        border-radius: 12px;
        font-weight: 600;
        padding: 0.75rem 2rem;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        border: none;
        position: relative;
        overflow: hidden;
    }
    .stButton > button::before {
        content: '';
        position: absolute;
        top: 0;
        left: -100%;
        width: 100%;
        height: 100%;
        background: linear-gradient(90deg, transparent, rgba(255,255,255,0.2), transparent);
        transition: left 0.5s;
    }
    .stButton > button:hover::before {
        left: 100%;
    }
    .stButton > button[kind="primary"] {
        background: var(--gradient-primary);
        color: white;
        box-shadow: 0 4px 20px rgba(59, 130, 246, 0.3);
    }
    .stButton > button[kind="primary"]:hover {
        background: linear-gradient(135deg, var(--primary-dark), var(--secondary-dark));
        transform: translateY(-2px);
        box-shadow: 0 8px 30px rgba(59, 130, 246, 0.4);
    }
    .stButton > button[kind="secondary"] {
        background: linear-gradient(135deg, var(--bg-primary), var(--bg-secondary));
        color: var(--text-secondary);
        border: 1px solid rgba(226, 232, 240, 0.8);
        backdrop-filter: blur(10px);
    }
    .stButton > button[kind="secondary"]:hover {
        background: linear-gradient(135deg, rgba(59, 130, 246, 0.1), rgba(139, 92, 246, 0.1));
        color: var(--primary-color);
        border-color: var(--primary-color);
        transform: translateY(-2px);
    }
    
    /* 进度条美化 */
    .stProgress > div > div > div {
        background: var(--gradient-primary);
        border-radius: 10px;
    }
    
    /* 选择框美化 */
    .stSelectbox > div > div {
        border-radius: 12px;
        border: 1px solid rgba(226, 232, 240, 0.8);
        background: rgba(255, 255, 255, 0.9);
        backdrop-filter: blur(10px);
    }
    
    /* 图标样式优化 */
    .icon-container {
        display: inline-flex;
        align-items: center;
        justify-content: center;
        width: 2.5rem;
        height: 2.5rem;
        border-radius: 50%;
        background: var(--gradient-primary);
        color: white;
        font-size: 1.2rem;
        margin-right: 0.75rem;
        box-shadow: var(--shadow-md);
        transition: all 0.3s ease;
    }
    .icon-container:hover {
        transform: scale(1.1);
        box-shadow: var(--shadow-lg);
    }
    
    /* 状态徽章样式 */
    .status-badge {
        display: inline-flex;
        align-items: center;
        padding: 0.25rem 0.75rem;
        border-radius: 9999px;
        font-size: 0.875rem;
        font-weight: 500;
        text-transform: uppercase;
        letter-spacing: 0.025em;
    }
    .status-badge.success {
        background: linear-gradient(135deg, rgba(16, 185, 129, 0.1), rgba(5, 150, 105, 0.1));
        color: var(--success-color);
        border: 1px solid rgba(16, 185, 129, 0.2);
    }
    .status-badge.warning {
        background: linear-gradient(135deg, rgba(245, 158, 11, 0.1), rgba(217, 119, 6, 0.1));
        color: var(--warning-color);
        border: 1px solid rgba(245, 158, 11, 0.2);
    }
    .status-badge.error {
        background: linear-gradient(135deg, rgba(239, 68, 68, 0.1), rgba(220, 38, 38, 0.1));
        color: var(--error-color);
        border: 1px solid rgba(239, 68, 68, 0.2);
    }
    .status-badge.info {
        background: linear-gradient(135deg, rgba(59, 130, 246, 0.1), rgba(37, 99, 235, 0.1));
        color: var(--info-color);
        border: 1px solid rgba(59, 130, 246, 0.2);
    }
    
    /* 响应式设计和移动端适配 */
    @media (max-width: 768px) {
        .main {
            padding: 0.5rem 1rem;
        }
        
        /* 移动端标签页优化 */
        .stTabs [data-baseweb="tab-list"] {
            flex-wrap: wrap;
            gap: 0.25rem;
            padding: 0.25rem;
        }
        .stTabs [data-baseweb="tab"] {
            height: 2.5rem;
            padding: 0 1rem;
            font-size: 14px;
            min-width: auto;
            flex: 1;
        }
        
        /* 移动端文件上传器 */
        .stFileUploader {
            padding: 2rem 1rem;
            margin: 0.5rem 0;
        }
        
        /* 移动端按钮 */
        .stButton > button {
            width: 100%;
            padding: 0.75rem 1rem;
            font-size: 14px;
        }
        
        /* 移动端指标容器 */
        .metric-container {
            padding: 1rem;
            margin: 0.5rem 0;
        }
        
        /* 移动端数据表格 */
        .stDataFrame {
            font-size: 12px;
        }
        
        /* 移动端标题 */
        h1 {
            font-size: 1.8rem !important;
        }
        h2 {
            font-size: 1.4rem !important;
        }
        h3 {
            font-size: 1.2rem !important;
        }
        
        /* 移动端图标容器 */
        .icon-container {
            width: 2rem;
            height: 2rem;
            font-size: 1rem;
            margin-right: 0.5rem;
        }
        
        /* 移动端状态徽章 */
        .status-badge {
            font-size: 0.75rem;
            padding: 0.2rem 0.5rem;
        }
    }
    
    @media (max-width: 480px) {
        .main {
            padding: 0.25rem 0.5rem;
        }
        
        /* 超小屏幕标签页 */
        .stTabs [data-baseweb="tab"] {
            height: 2.25rem;
            padding: 0 0.75rem;
            font-size: 12px;
        }
        
        /* 超小屏幕文件上传器 */
        .stFileUploader {
            padding: 1.5rem 0.75rem;
        }
        
        /* 超小屏幕标题 */
        h1 {
            font-size: 1.5rem !important;
        }
        h2 {
            font-size: 1.2rem !important;
        }
        h3 {
            font-size: 1rem !important;
        }
        
        /* 超小屏幕指标容器 */
        .metric-container {
            padding: 0.75rem;
        }
    }
    
    /* 平板端适配 */
    @media (min-width: 769px) and (max-width: 1024px) {
        .main {
            padding: 1rem 1.5rem;
        }
        
        .stTabs [data-baseweb="tab"] {
            padding: 0 1.5rem;
        }
        
        .metric-container {
            padding: 1.5rem;
        }
    }
    
    /* 大屏幕优化 */
    @media (min-width: 1400px) {
        .main {
            max-width: 1600px;
        }
        
        .stTabs [data-baseweb="tab"] {
            padding: 0 2.5rem;
            height: 4rem;
            font-size: 16px;
        }
        
        .metric-container {
            padding: 2.5rem;
        }
    }
    
    /* 触摸设备优化 */
    @media (hover: none) and (pointer: coarse) {
        .stButton > button {
            min-height: 44px;
            padding: 0.75rem 1.5rem;
        }
        
        .stTabs [data-baseweb="tab"] {
            min-height: 44px;
        }
        
        .icon-container {
            min-width: 44px;
            min-height: 44px;
        }
    }
    </style>
    """, unsafe_allow_html=True)
    
    # 现代化标题设计
    st.markdown("""
    <div style='text-align: center; margin-bottom: 3rem; padding: 2rem 0;'>
        <div style='background: linear-gradient(135deg, rgba(255,255,255,0.9), rgba(248,250,252,0.9)); 
                    backdrop-filter: blur(15px); 
                    border-radius: 20px; 
                    padding: 2rem; 
                    box-shadow: 0 10px 40px rgba(0,0,0,0.1); 
                    border: 1px solid rgba(255,255,255,0.2); 
                    display: inline-block; 
                    position: relative; 
                    overflow: hidden;'>
            <div style='position: absolute; top: 0; left: 0; right: 0; bottom: 0; 
                        background: linear-gradient(45deg, rgba(102,126,234,0.1), rgba(118,75,162,0.1)); 
                        animation: shimmer 3s ease-in-out infinite;'></div>
            <h1 style='background: linear-gradient(135deg, #667eea, #764ba2); 
                       -webkit-background-clip: text; 
                       -webkit-text-fill-color: transparent; 
                       background-clip: text; 
                       font-weight: 800; 
                       font-size: 2.5rem; 
                       margin: 0; 
                       position: relative; 
                       z-index: 1; 
                       text-shadow: 0 2px 4px rgba(0,0,0,0.1);'>
                📊 鹭府预定匹配工具 v2.0
            </h1>
            <p style='color: #64748b; 
                      font-size: 1.1rem; 
                      margin: 0.5rem 0 0 0; 
                      position: relative; 
                      z-index: 1; 
                      font-weight: 500;'>
                智能数据匹配 • 高效预订管理 • 可视化分析
            </p>
        </div>
    </div>
    <style>
    @keyframes shimmer {
        0% { transform: translateX(-100%) rotate(45deg); }
        50% { transform: translateX(100%) rotate(45deg); }
        100% { transform: translateX(-100%) rotate(45deg); }
    }
    </style>
    """, unsafe_allow_html=True)
    
    # 初始化应用
    if 'app' not in st.session_state:
        st.session_state.app = ReservationMatcherWeb()
    
    app = st.session_state.app
    
    # 三个主要标签页
    tab1, tab2, tab3 = st.tabs(["📁 文件处理", "📊 结果查看", "📈 数据分析"])
    
    # 标签页内容
    with tab1:
        # 文件上传和数据匹配合并
        col1, col2 = st.columns([1, 1])
        
        with col1:
            st.subheader("📤 文件上传")
            app.load_files()
            
        with col2:
            st.subheader("⚡ 数据匹配")
            
            # 验证文件
            is_valid, message = app.validate_files()
            
            if not is_valid:
                # 现代化警告提示
                st.markdown(f"""
                <div style='background: linear-gradient(135deg, rgba(245, 158, 11, 0.1), rgba(217, 119, 6, 0.1)); 
                            border-radius: 12px; 
                            padding: 1rem 1.5rem; 
                            border-left: 4px solid #f59e0b; 
                            margin: 1rem 0;'>
                    <div style='display: flex; align-items: center; gap: 0.5rem;'>
                        <span style='font-size: 1.2rem;'>⚠️</span>
                        <strong style='color: #92400e;'>等待文件上传</strong>
                    </div>
                    <p style='margin: 0.5rem 0 0 0; color: #78350f; font-size: 0.9rem;'>
                        {message}
                    </p>
                </div>
                """, unsafe_allow_html=True)
            else:
                # 现代化成功提示
                st.markdown("""
                <div style='background: linear-gradient(135deg, rgba(16, 185, 129, 0.1), rgba(5, 150, 105, 0.1)); 
                            border-radius: 12px; 
                            padding: 1rem 1.5rem; 
                            border-left: 4px solid #10b981; 
                            margin: 1rem 0;'>
                    <div style='display: flex; align-items: center; gap: 0.5rem;'>
                        <span style='font-size: 1.2rem;'>✅</span>
                        <strong style='color: #059669;'>文件已就绪</strong>
                    </div>
                    <p style='margin: 0.5rem 0 0 0; color: #064e3b; font-size: 0.9rem;'>
                        所有文件已成功加载，可以开始数据匹配
                    </p>
                </div>
                """, unsafe_allow_html=True)
                
                if st.button("🚀 开始智能匹配", type="primary", use_container_width=True):
                    # 现代化进度指示器
                    progress_container = st.container()
                    with progress_container:
                        st.markdown("""
                        <div style='background: linear-gradient(135deg, rgba(59, 130, 246, 0.1), rgba(139, 92, 246, 0.1)); 
                                    border-radius: 12px; 
                                    padding: 1.5rem; 
                                    margin: 1rem 0; 
                                    text-align: center;'>
                            <div style='color: #3b82f6; font-weight: 600; margin-bottom: 1rem;'>
                                🔄 正在进行智能数据匹配...
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        # 进度条
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        
                        # 模拟匹配过程的进度更新
                        import time
                        
                        status_text.text("📊 正在分析美团数据...")
                        progress_bar.progress(20)
                        time.sleep(0.5)
                        
                        status_text.text("📋 正在处理预订记录...")
                        progress_bar.progress(40)
                        time.sleep(0.5)
                        
                        status_text.text("🔍 正在执行智能匹配算法...")
                        progress_bar.progress(70)
                        time.sleep(0.5)
                        
                        status_text.text("✨ 正在生成匹配结果...")
                        progress_bar.progress(90)
                        
                        # 执行实际匹配
                        success, result_message = app.match_data()
                        
                        progress_bar.progress(100)
                        status_text.text("🎉 匹配完成！")
                        time.sleep(0.5)
                        
                        # 清除进度指示器
                        progress_container.empty()
                        
                    if success:
                        # 现代化成功提示
                        st.markdown(f"""
                        <div style='background: linear-gradient(135deg, rgba(16, 185, 129, 0.1), rgba(5, 150, 105, 0.1)); 
                                    border-radius: 12px; 
                                    padding: 1rem 1.5rem; 
                                    border-left: 4px solid #10b981; 
                                    margin: 1rem 0;'>
                            <div style='display: flex; align-items: center; gap: 0.5rem;'>
                                <span style='font-size: 1.2rem;'>🎉</span>
                                <strong style='color: #059669;'>匹配成功完成！</strong>
                            </div>
                            <p style='margin: 0.5rem 0 0 0; color: #064e3b; font-size: 0.9rem;'>
                                {result_message}
                            </p>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        # 导航提示
                        st.markdown("""
                        <div style='background: linear-gradient(135deg, rgba(59, 130, 246, 0.1), rgba(139, 92, 246, 0.1)); 
                                    border-radius: 12px; 
                                    padding: 1rem 1.5rem; 
                                    border-left: 4px solid #3b82f6; 
                                    margin: 1rem 0; 
                                    text-align: center;'>
                            <p style='margin: 0; color: #1e40af; font-weight: 500;'>
                                💡 请切换到 <strong>"📊 结果查看"</strong> 标签页查看匹配结果
                            </p>
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        # 现代化错误提示
                        st.markdown(f"""
                        <div style='background: linear-gradient(135deg, rgba(239, 68, 68, 0.1), rgba(220, 38, 38, 0.1)); 
                                    border-radius: 12px; 
                                    padding: 1rem 1.5rem; 
                                    border-left: 4px solid #ef4444; 
                                    margin: 1rem 0;'>
                            <div style='display: flex; align-items: center; gap: 0.5rem;'>
                                <span style='font-size: 1.2rem;'>❌</span>
                                <strong style='color: #dc2626;'>匹配失败</strong>
                            </div>
                            <p style='margin: 0.5rem 0 0 0; color: #7f1d1d; font-size: 0.9rem;'>
                                {result_message}
                            </p>
                        </div>
                        """, unsafe_allow_html=True)
    
    with tab2:
        # 查看结果和导出合并
        col1, col2 = st.columns([3, 1])
        
        with col1:
            app.display_results()
            
        with col2:
            # 现代化导出区域
            st.markdown("""
            <div style='background: linear-gradient(135deg, rgba(34, 197, 94, 0.1), rgba(22, 163, 74, 0.1)); 
                        border-radius: 12px; 
                        padding: 1.5rem; 
                        margin: 1rem 0; 
                        border: 1px solid rgba(34, 197, 94, 0.2);'>
                <h4 style='color: #15803d; margin: 0 0 1rem 0; font-weight: 600; display: flex; align-items: center; gap: 0.5rem;'>
                    📥 数据导出中心
                </h4>
                <p style='color: #166534; font-size: 0.9rem; margin: 0;'>
                    将匹配结果导出为Excel文件，便于后续处理和分析
                </p>
            </div>
            """, unsafe_allow_html=True)
            app.export_results()
    
    with tab3:
        # 数据分析标签页
        app.show_data_analysis()
    


if __name__ == "__main__":
    main()